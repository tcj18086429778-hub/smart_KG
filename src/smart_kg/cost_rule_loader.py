"""成本规则抽取器。

本文件负责从最新版 Excel 规则表中抽取结构化成本规则，并统一输出为
`CostRuleEntry` 列表或 JSON 文件，是“Excel 规则 -> 图谱规则/文件规则”的共同入口。

在整体链路中的位置：
1. 读取 Excel 原始规则。
2. 解析顶层规则表和命名成本块。
3. 输出可直接用于 GPKG 成本化和 Neo4j 入库的标准规则。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from .data_loader import read_json, write_json
from .logging_utils import instrument_module_functions


VOLTAGE_LEVELS = ["10kV", "35kV", "110kV", "220kV", "330kV", "500kV", "750kV", "±800kV", "1000kV", "±1100kV"]
IGNORED_VALUES = {"", "\\", "/", "不参与计算", "无", "none", "None", "NULL", "null"}
NEGOTIABLE_VALUES = {"?", "？", "-1", "单独计列", "需确认", "TODO", "？TODO？"}
logger = logging.getLogger(__name__)


class CostRuleEntry(BaseModel):
    """标准化成本规则条目。

    字段涵盖规则标识、来源表、要素分类、计算模式、效果值、
    电压等级、缓冲距离、避让模式及匹配条件等完整业务信息。
    """
    rule_id: str
    rule_name: str
    rule_set_version: str | None = None
    source_table: str
    source_row: int
    feature_type_code: str | None = None
    feature_type_name: str | None = None
    feature_subtype_code: str | None = None
    feature_subtype_name: str | None = None
    feature_level: str | None = None
    geometry_kind: str | None = None
    calc_mode: str
    effect_target: str = "ALL"
    effect_value: float | None = None
    effect_value_status: str
    effect_attr: str | None = None
    effect_unit: str | None = None
    voltage_level: str | None = None
    buffer_distance_m: float | None = None
    avoidance_mode: str | None = None
    raw_value: Any = None
    reason_code: int
    priority: int = 100
    match_condition_json: dict[str, Any] = Field(default_factory=dict)
    enabled: bool = True


def standardize_cost_rules(excel_path: Path, out_path: Path | None = None, rule_set_version: str | None = None) -> list[dict[str, Any]]:
    """标准化 Excel 成本规则为 JSON 字典列表。

    参数：
        excel_path: 成本分类体系 Excel 文件路径。
        out_path: 可选输出 JSON 路径。
        rule_set_version: 规则集版本标签。

    返回：
        规则字典列表。
    """
    logger.info("开始标准化成本规则：excel_path=%s, out_path=%s, rule_set_version=%s", excel_path, out_path, rule_set_version)
    entries = extract_cost_rules(excel_path, rule_set_version=rule_set_version)
    rows = [entry.model_dump(mode="json") for entry in entries]
    if out_path:
        write_json(out_path, rows)
        logger.info("成本规则已写出：out_path=%s, count=%s", out_path, len(rows))
    return rows


def load_cost_rules(path: Path) -> list[CostRuleEntry]:
    """从 JSON 文件加载 CostRuleEntry 对象列表。

    参数：
        path: JSON 文件路径。

    返回：
        CostRuleEntry 对象列表。
    """
    logger.info("加载成本规则文件：path=%s", path)
    return [CostRuleEntry.model_validate(row) for row in read_json(path)]


def extract_cost_rules(excel_path: Path, rule_set_version: str | None = None) -> list[CostRuleEntry]:
    """从 Excel 中提取全部成本规则。

    遍历工作簿的所有工作表，解析顶层规则表和命名成本块，
    为缺失的原因编码自动分配回退值。

    参数：
        excel_path: 成本分类体系 Excel 路径。
        rule_set_version: 版本标签。

    返回：
        CostRuleEntry 对象列表。
    """
    logger.info("开始解析 Excel 成本规则：excel_path=%s, rule_set_version=%s", excel_path, rule_set_version)
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    entries: list[CostRuleEntry] = []
    seen: set[tuple[Any, ...]] = set()
    for ws in wb.worksheets:
        if ws.title == "选线要素分类及代码":
            continue
        logger.debug("处理工作表：sheet=%s", ws.title)
        header_row = find_header_row(ws)
        if header_row is None:
            logger.debug("工作表缺少标准头部，跳过：sheet=%s", ws.title)
            continue
        header = read_header(ws, header_row)
        logger.debug("已识别头部：sheet=%s, header=%s", ws.title, header)
        entries.extend(extract_top_table_rules(ws, header, header_row, rule_set_version, seen))
        entries.extend(extract_named_cost_blocks(ws, rule_set_version, seen))
    for index, entry in enumerate(entries, start=1):
        if entry.reason_code <= 0:
            entry.reason_code = 1000 + index if entry.calc_mode == "FORBIDDEN" else 2000 + index
    logger.info("Excel 成本规则解析完成：excel_path=%s, count=%s", excel_path, len(entries))
    return entries


def find_header_row(ws: Any) -> int | None:
    """在工作表前 10 行中定位表头行。

    通过扫描"代码"+"因素类型"或"要素中类"特征列来识别表头。

    参数：
        ws: openpyxl 工作表对象。

    返回：
        表头所在行号，未找到时返回 None。
    """
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [normalize_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if "代码" in values and ("因素类型" in values or "要素中类" in values):
            return row_idx
    return None


def read_header(ws: Any, row_idx: int) -> dict[str, int]:
    """读取表头行，返回列名到列索引的映射。

    参数：
        ws: openpyxl 工作表。
        row_idx: 表头所在行号。

    返回：
        {列名: 列号} 字典。
    """
    header: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = normalize_text(ws.cell(row_idx, col).value)
        if value and value not in header:
            header[value] = col
    return header


def extract_top_table_rules(
    ws: Any,
    header: dict[str, int],
    header_row: int,
    rule_set_version: str | None,
    seen: set[tuple[Any, ...]],
) -> list[CostRuleEntry]:
    """从工作表的顶层规则表格中提取禁建规则和成本规则。

    遍历表头下方的数据行，读取要素名称、代码、避让模式、成本值和
    电压等级等字段，生成相应的 CostRuleEntry。

    参数：
        ws: 工作表对象。
        header: 表头列映射。
        header_row: 表头行号。
        rule_set_version: 规则集版本。
        seen: 去重集合。

    返回：
        CostRuleEntry 列表。
    """
    entries: list[CostRuleEntry] = []
    factor_col = header.get("因素类型") or header.get("要素中类")
    code_col = header.get("代码")
    if not factor_col or not code_col:
        logger.debug("顶层规则表缺少关键列，跳过：sheet=%s, header=%s", ws.title, header)
        return entries
    second_col = header.get("二层因素") or header.get("要素大类")
    direct_col = first_header_containing(header, "是否直接避让")
    buffer_flag_col = first_header_containing(header, "是否按缓冲距离直接避让")
    buffer_col = first_header_exact(header, "缓冲距离")
    cost_col = first_header_exact(header, "成本值") or first_header_exact(header, "成本系数")
    voltage_cols = {voltage: header.get(f"{voltage}成本") for voltage in VOLTAGE_LEVELS if header.get(f"{voltage}成本")}
    second_factor: str | None = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        feature_name = normalize_text(ws.cell(row_idx, factor_col).value)
        code = normalize_code(ws.cell(row_idx, code_col).value)
        current_second = normalize_text(ws.cell(row_idx, second_col).value) if second_col else None
        second_factor = current_second or second_factor
        if not feature_name or not code:
            continue
        logger.debug("解析顶层规则行：sheet=%s, row=%s, feature_name=%s, code=%s", ws.title, row_idx, feature_name, code)
        geometry_kind = geometry_kind_from_row(ws, header, row_idx)
        buffer_distance = parse_buffer_distance(ws.cell(row_idx, buffer_col).value if buffer_col else None)
        direct_avoid = is_yes(ws.cell(row_idx, direct_col).value if direct_col else None)
        buffer_avoid = is_yes(ws.cell(row_idx, buffer_flag_col).value if buffer_flag_col else None)
        avoidance_mode = None
        if buffer_avoid and buffer_distance is not None:
            avoidance_mode = "BUFFER"
        elif direct_avoid:
            avoidance_mode = "DIRECT"
        if avoidance_mode:
            logger.debug(
                "写入禁建规则：sheet=%s, row=%s, feature_name=%s, code=%s, avoidance_mode=%s, buffer_distance=%s",
                ws.title,
                row_idx,
                feature_name,
                code,
                avoidance_mode,
                buffer_distance,
            )
            entries.append(
                make_rule(
                    ws=ws,
                    row_idx=row_idx,
                    feature_name=feature_name,
                    code=code,
                    second_factor=second_factor,
                    geometry_kind=geometry_kind,
                    calc_mode="FORBIDDEN",
                    effect_value_status="FORBIDDEN",
                    effect_value=None,
                    raw_value="direct_or_buffer_avoid",
                    effect_attr=None,
                    voltage_level=None,
                    buffer_distance_m=buffer_distance,
                    avoidance_mode=avoidance_mode,
                    rule_set_version=rule_set_version,
                    seen=seen,
                    priority=1000,
                )
            )
        raw_values: list[tuple[str | None, Any]] = []
        if cost_col:
            raw_values.append((None, ws.cell(row_idx, cost_col).value))
        raw_values.extend((voltage, ws.cell(row_idx, col).value) for voltage, col in voltage_cols.items())
        for voltage, raw_value in raw_values:
            value_status, effect_value = normalize_effect_value(raw_value)
            if value_status == "NOT_CONSIDERED":
                continue
            calc_mode = "FORBIDDEN" if value_status == "FORBIDDEN" else calc_mode_for_header(header, raw_value)
            logger.debug(
                "写入成本规则：sheet=%s, row=%s, feature_name=%s, code=%s, voltage=%s, raw_value=%s, calc_mode=%s, effect_value=%s",
                ws.title,
                row_idx,
                feature_name,
                code,
                voltage,
                raw_value,
                calc_mode,
                effect_value,
            )
            entries.append(
                make_rule(
                    ws=ws,
                    row_idx=row_idx,
                    feature_name=feature_name,
                    code=code,
                    second_factor=second_factor,
                    geometry_kind=geometry_kind,
                    calc_mode=calc_mode,
                    effect_value_status=value_status,
                    effect_value=effect_value,
                    raw_value=raw_value,
                    effect_attr=effect_attr_for_value(raw_value, geometry_kind, calc_mode),
                    voltage_level=voltage,
                    buffer_distance_m=buffer_distance,
                    avoidance_mode="DIRECT" if calc_mode == "FORBIDDEN" else None,
                    rule_set_version=rule_set_version,
                    seen=seen,
                    priority=1000 if calc_mode == "FORBIDDEN" else 200,
                )
            )
    return entries


def extract_named_cost_blocks(ws: Any, rule_set_version: str | None, seen: set[tuple[Any, ...]]) -> list[CostRuleEntry]:
    """从工作表中提取命名成本块规则。

    扫描工作表中以"因素类型"/"要素种类"为标记的成本块，
    提取块内按要素名称和成本值/成本系数定义的成本规则。

    参数：
        ws: 工作表对象。
        rule_set_version: 规则集版本。
        seen: 去重集合。

    返回：
        CostRuleEntry 列表。
    """
    entries: list[CostRuleEntry] = []
    for row_idx in range(1, ws.max_row + 1):
        values = [normalize_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        for col_idx, value in enumerate(values, start=1):
            if value not in {"因素类型", "要素种类", "要素中类"}:
                continue
            cost_col = col_idx + 1
            if normalize_text(ws.cell(row_idx, cost_col).value) not in {"成本值", "成本系数"}:
                continue
            logger.debug("识别到命名成本块：sheet=%s, header_row=%s, feature_col=%s, cost_col=%s", ws.title, row_idx, col_idx, cost_col)
            for data_row in range(row_idx + 1, ws.max_row + 1):
                feature_name = normalize_text(ws.cell(data_row, col_idx).value)
                raw_value = ws.cell(data_row, cost_col).value
                if not feature_name:
                    break
                value_status, effect_value = normalize_effect_value(raw_value)
                if value_status == "NOT_CONSIDERED":
                    continue
                calc_mode = "FORBIDDEN" if value_status == "FORBIDDEN" else ("MAIN_COST_SCALING" if normalize_text(ws.cell(row_idx, cost_col).value) == "成本系数" else "MAIN_COST_INCREMENT")
                logger.debug(
                    "写入命名成本块规则：sheet=%s, row=%s, feature_name=%s, raw_value=%s, calc_mode=%s, effect_value=%s",
                    ws.title,
                    data_row,
                    feature_name,
                    raw_value,
                    calc_mode,
                    effect_value,
                )
                entries.append(
                    make_rule(
                        ws=ws,
                        row_idx=data_row,
                        feature_name=feature_name,
                        code=None,
                        second_factor=None,
                        geometry_kind=None,
                        calc_mode=calc_mode,
                        effect_value_status=value_status,
                        effect_value=effect_value,
                        raw_value=raw_value,
                        effect_attr=effect_attr_for_value(raw_value, None, calc_mode),
                        voltage_level=None,
                        buffer_distance_m=None,
                        avoidance_mode="DIRECT" if calc_mode == "FORBIDDEN" else None,
                        rule_set_version=rule_set_version,
                        seen=seen,
                        priority=900 if calc_mode == "FORBIDDEN" else 300,
                    )
                )
    return entries


def make_rule(
    ws: Any,
    row_idx: int,
    feature_name: str,
    code: str | None,
    second_factor: str | None,
    geometry_kind: str | None,
    calc_mode: str,
    effect_value_status: str,
    effect_value: float | None,
    raw_value: Any,
    effect_attr: str | None,
    voltage_level: str | None,
    buffer_distance_m: float | None,
    avoidance_mode: str | None,
    rule_set_version: str | None,
    seen: set[tuple[Any, ...]],
    priority: int,
) -> CostRuleEntry:
    """构造单个 CostRuleEntry 对象。

    生成唯一 rule_id、构造匹配条件 JSON 并检查去重。

    参数：
        ws: 源工作表。
        row_idx: 源行号。
        feature_name: 要素名称。
        code: 要素子类编码。
        second_factor: 二层因素/要素大类。
        geometry_kind: 几何类型（point/line/polygon）。
        calc_mode: 计算模式。
        effect_value_status: 效果值状态。
        effect_value: 效果数值。
        raw_value: Excel 原始值。
        effect_attr: 效果属性。
        voltage_level: 电压等级。
        buffer_distance_m: 缓冲距离（米）。
        avoidance_mode: 避让模式。
        rule_set_version: 规则集版本。
        seen: 去重集合。
        priority: 优先级。

    返回：
        构造好的 CostRuleEntry 对象。

    抛出：
        DuplicateRule: 当规则键重复时。
    """
    subtype_code = code
    type_code = f"{code[:2]}00" if code and len(code) >= 2 else None
    type_name = second_factor
    key = (ws.title, row_idx, feature_name, subtype_code, calc_mode, effect_value_status, voltage_level, normalize_text(raw_value))
    if key in seen:
        raise DuplicateRule
    seen.add(key)
    raw_slug = slug(normalize_text(raw_value) or effect_value_status)
    voltage_slug = slug(voltage_level or "all")
    code_slug = subtype_code or slug(feature_name)
    rule_id = f"cost_rule:{slug(ws.title)}:{code_slug}:{voltage_slug}:{row_idx}:{raw_slug}"
    if calc_mode == "FORBIDDEN":
        rule_id = f"constraint_rule:{slug(ws.title)}:{code_slug}:{voltage_slug}:{row_idx}"
    condition = {"field": "feature_subtype_name", "operator": "eq", "value": feature_name}
    if subtype_code:
        condition = {"field": "feature_subtype_code", "operator": "eq", "value": subtype_code}
    if voltage_level:
        condition = {"logic": "AND", "conditions": [condition, {"field": "voltage_level", "operator": "eq", "value": voltage_level}]}
    logger.debug(
        "构造规则对象：sheet=%s, row=%s, rule_id=%s, feature_name=%s, subtype_code=%s, calc_mode=%s, voltage_level=%s, effect_value=%s",
        ws.title,
        row_idx,
        rule_id,
        feature_name,
        subtype_code,
        calc_mode,
        voltage_level,
        effect_value,
    )
    return CostRuleEntry(
        rule_id=rule_id,
        rule_name=f"{feature_name}{voltage_level or ''}{'禁建' if calc_mode == 'FORBIDDEN' else '成本'}",
        rule_set_version=rule_set_version,
        source_table=ws.title,
        source_row=row_idx,
        feature_type_code=type_code,
        feature_type_name=type_name,
        feature_subtype_code=subtype_code,
        feature_subtype_name=feature_name,
        geometry_kind=geometry_kind,
        calc_mode=calc_mode,
        effect_target="ALL",
        effect_value=effect_value,
        effect_value_status=effect_value_status,
        effect_attr=effect_attr,
        effect_unit=effect_unit_for_attr(effect_attr, calc_mode),
        voltage_level=voltage_level,
        buffer_distance_m=buffer_distance_m,
        avoidance_mode=avoidance_mode,
        raw_value=raw_value,
        reason_code=0,
        priority=priority,
        match_condition_json=condition,
        enabled=True,
    )


class DuplicateRule(Exception):
    """规则去重异常，当同一工作表中存在完全相同的规则键时触发。"""
    pass


def normalize_text(value: Any) -> str | None:
    """把任意值转为去空白字符串，空字符串返回 None。

    参数：
        value: 任意输入值。

    返回：
        标准化后的非空字符串，或 None。
    """
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_code(value: Any) -> str | None:
    """将单元格值标准化为四位数字编码字符串。

    参数：
        value: Excel 单元格值。

    返回：
        四位编码字符串或 None（非标准编码）。
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if re.fullmatch(r"\d+(\.0)?", text):
        return str(int(float(text)))
    return text if re.fullmatch(r"\d{4}", text) else None


def first_header_containing(header: dict[str, int], text: str) -> int | None:
    """返回第一个包含指定文本的列名对应的列索引。

    参数：
        header: 列名到列索引的映射。
        text: 要搜索的子串。

    返回：
        列索引或 None。
    """
    for key, col in header.items():
        if text in key:
            return col
    return None


def first_header_exact(header: dict[str, int], text: str) -> int | None:
    """精确匹配列名并返回对应列索引。

    参数：
        header: 列名到列索引的映射。
        text: 要匹配的精确列名。

    返回：
        列索引或 None。
    """
    return header.get(text)


def geometry_kind_from_row(ws: Any, header: dict[str, int], row_idx: int) -> str | None:
    """根据行中"点"/"线"/"面"列的勾选标记推断几何类型。

    参数：
        ws: 工作表。
        header: 表头列映射。
        row_idx: 数据行号。

    返回：
        "point"、"line"、"polygon" 或 None。
    """
    for title, kind in [("点", "point"), ("线", "line"), ("面", "polygon")]:
        col = header.get(title)
        if col and normalize_text(ws.cell(row_idx, col).value) == "√":
            return kind
    return None


def is_yes(value: Any) -> bool:
    """判断单元格值是否为"是"。

    参数：
        value: 单元格值。

    返回：
        True 当值为"是"。
    """
    return normalize_text(value) == "是"


def parse_buffer_distance(value: Any) -> float | None:
    """从单元格值中解析缓冲距离数值。

    参数：
        value: 单元格值。

    返回：
        提取后的最大数值（米），无法解析时返回 None。
    """
    text = normalize_text(value)
    if not text or text in IGNORED_VALUES:
        return None
    nums = parse_numbers(text)
    return max(nums) if nums else None


def normalize_effect_value(value: Any) -> tuple[str, float | None]:
    """将单元格值标准化为效果值状态和数值。

    区分 NOT_CONSIDERED、NEGOTIABLE、FORBIDDEN（max）和 NUMERIC。

    参数：
        value: 单元格值。

    返回：
        (状态字符串, 数值或 None) 元组。
    """
    text = normalize_text(value)
    if text is None or text in IGNORED_VALUES:
        return "NOT_CONSIDERED", None
    if text.lower() == "max":
        return "FORBIDDEN", None
    if any(token in text for token in NEGOTIABLE_VALUES):
        return "NEGOTIABLE", None
    if isinstance(value, (int, float)):
        return "NUMERIC", float(value)
    nums = parse_numbers(text)
    if nums:
        return "NUMERIC", max(nums)
    return "NEGOTIABLE", None


def parse_numbers(text: str) -> list[float]:
    """从文本中提取所有数值。

    处理 "00xxx" 前缀的小数（如 003 -> 0.03）和常规浮点数。

    参数：
        text: 待解析文本。

    返回：
        提取的浮点数列表。
    """
    values: list[float] = []
    for match in re.finditer(r"\d+(?:\.\d+)?", text):
        token = match.group(0)
        if token.startswith("00") and "." not in token:
            values.append(float(f"0.{token.lstrip('0') or '0'}"))
        else:
            values.append(float(token))
    return values


def calc_mode_for_header(header: dict[str, int], raw_value: Any) -> str:
    """根据表头列名和原始值判断计算模式。

    参数：
        header: 表头列映射。
        raw_value: 原始单元格值。

    返回：
        "MAIN_COST_SCALING" 或 "MAIN_COST_INCREMENT"。
    """
    raw_text = normalize_text(raw_value) or ""
    if "系数" in raw_text:
        return "MAIN_COST_SCALING"
    if "成本系数" in header:
        return "MAIN_COST_SCALING"
    return "MAIN_COST_INCREMENT"


def effect_attr_for_value(raw_value: Any, geometry_kind: str | None, calc_mode: str) -> str | None:
    """根据原始值和几何类型推断效果属性（面积/长度/计数）。

    禁建和系数模式不需要效果属性，返回 None。

    参数：
        raw_value: 原始值。
        geometry_kind: 几何类型。
        calc_mode: 计算模式。

    返回：
        "S_AREA"/"S_LTH"/"S_CNT" 或 None。
    """
    if calc_mode in {"FORBIDDEN", "MAIN_COST_SCALING"}:
        return None
    text = normalize_text(raw_value) or ""
    if any(unit in text for unit in ["㎡", "平方米", "亩"]):
        return "S_AREA"
    if any(unit in text for unit in ["公里", "km", "千米"]):
        return "S_LTH"
    if any(unit in text for unit in ["户", "座", "棵", "根", "个", "处", "池", "次"]):
        return "S_CNT"
    if geometry_kind == "line":
        return "S_LTH"
    if geometry_kind == "polygon":
        return "S_AREA"
    return "S_CNT"


def effect_unit_for_attr(effect_attr: str | None, calc_mode: str) -> str | None:
    """根据效果属性和计算模式返回对应的计量单位。

    参数：
        effect_attr: 效果属性。
        calc_mode: 计算模式。

    返回：
        单位字符串（如"万元/公里"、"万元/处"等）。
    """
    if calc_mode == "MAIN_COST_SCALING":
        return "系数"
    if effect_attr == "S_AREA":
        return "万元/面积单位"
    if effect_attr == "S_LTH":
        return "万元/公里"
    if effect_attr == "S_CNT":
        return "万元/处"
    return "万元"


def slug(value: Any) -> str:
    """将任意值转为 URL 友好标识符。

    参数：
        value: 任意值。

    返回：
        下划线分隔的标识符字符串。
    """
    text = normalize_text(value) or "none"
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff+-]+", "_", text).strip("_")


instrument_module_functions(globals(), logger)
